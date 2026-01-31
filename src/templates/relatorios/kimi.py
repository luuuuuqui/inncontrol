import streamlit as st  # pyright: ignore[reportMissingImports]
import pandas as pd
from views import View
from datetime import datetime, timedelta, date
from decimal import Decimal
from collections import defaultdict
from io import BytesIO
import calendar

try:
    import plotly.express as px
    import plotly.graph_objects as go

    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib.enums import TA_CENTER

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    EXCEL_AVAILABLE = True
except ImportError:
    try:
        import xlsxwriter

        EXCEL_AVAILABLE = True
    except ImportError:
        EXCEL_AVAILABLE = False

try:
    from fpdf import FPDF

    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False


class RelatoriosKimiUI:
    """
    Módulo de Relatórios Administrativos - InnControl
    Dashboards interativos com Plotly, análise temporal e exportação profissional
    """

    @staticmethod
    def main():
        st.header("Relatórios Administrativos")
        st.markdown("Dashboard gerencial completo para análise de desempenho")

        filtro_periodo = RelatoriosKimiUI._render_filtros_avancados()

        tab1, tab2, tab3, tab4 = st.tabs(
            [
                "Visão Geral",
                "Análise Temporal",
                "Serviços & Adicionais",
                "Exportar Relatórios",
            ]
        )

        with tab1:
            RelatoriosKimiUI._render_visao_geral_plus(filtro_periodo)

        with tab2:
            RelatoriosKimiUI._render_comparativo_temporal(filtro_periodo)

        with tab3:
            RelatoriosKimiUI._render_servicos_adicionais(filtro_periodo)

        with tab4:
            RelatoriosKimiUI._render_exportacao_profissional(filtro_periodo)

    @staticmethod
    def _render_filtros_avancados():
        """Filtros estilo Claude (ano/mês) + seletor de período customizado"""
        st.sidebar.header("Filtros do Período")

        modo_filtro = st.sidebar.radio(
            "Modo de seleção:", ["Mês/Ano", "Período Customizado"], horizontal=True
        )

        if modo_filtro == "Mês/Ano":
            col1, col2 = st.sidebar.columns(2)
            with col1:
                ano = st.selectbox(
                    "Ano:",
                    options=list(
                        range(datetime.now().year - 2, datetime.now().year + 2)
                    ),
                    index=2,
                )
            with col2:
                mes = st.selectbox(
                    "Mês:",
                    options=list(range(1, 13)),
                    format_func=lambda x: calendar.month_name[x],
                    index=datetime.now().month - 1,
                )
            inicio = date(ano, mes, 1)
            fim = date(ano, mes, calendar.monthrange(ano, mes)[1])
        else:
            col1, col2 = st.sidebar.columns(2)
            with col1:
                inicio = st.date_input(
                    "Início", value=date.today().replace(day=1), format="DD/MM/YYYY"
                )
            with col2:
                fim = st.date_input("Fim", value=date.today(), format="DD/MM/YYYY")

        st.sidebar.markdown("---")
        cols_btn = st.sidebar.columns(3)
        if cols_btn[0].button("Este Mês", use_container_width=True):
            hoje = date.today()
            st.session_state["forcing_rerun"] = {
                "inicio": hoje.replace(day=1),
                "fim": hoje,
            }
        if cols_btn[1].button("Mês Ant.", use_container_width=True):
            hoje = date.today()
            primeiro_dia = hoje.replace(day=1)
            ultimo_mes = primeiro_dia - timedelta(days=1)
            st.session_state["forcing_rerun"] = {
                "inicio": ultimo_mes.replace(day=1),
                "fim": ultimo_mes,
            }

        return {"inicio": inicio, "fim": fim, "modo": modo_filtro}

    @staticmethod
    def _processar_dados_completos(periodo):
        """Processamento robusto dos dados (feature Claude: métricas agregadas)"""
        try:
            reservas = View.reserva_listar() or []
            pagamentos = View.pagamento_listar() or []
            consumos = View.consumo_listar() or []
            adicionais = View.adicional_listar() or []
            hospedes = View.hospede_listar() or []
            quartos = View.quarto_listar() or []

            reservas_periodo = []
            hospedes_unicos = set()
            reservas_por_status = defaultdict(int)
            reservas_por_tipo = defaultdict(int)

            for r in reservas:
                try:
                    checkin = datetime.strptime(r.get_data_checkin(), "%Y-%m-%d").date()
                    if periodo["inicio"] <= checkin <= periodo["fim"]:
                        reservas_periodo.append(r)
                        hospedes_unicos.add(r.get_id_hospede())
                        reservas_por_status[r.get_status()] += 1

                        quarto = View.quarto_listar_id(r.get_id_quarto())
                        if quarto:
                            tipo = View.tipoquarto_listar_id(
                                quarto.get_id_quarto_tipo()
                            )
                            if tipo:
                                reservas_por_tipo[tipo.get_nome()] += 1
                except:
                    continue

            faturamento = Decimal("0.00")
            for p in pagamentos:
                try:
                    data_pag = datetime.strptime(
                        p.get_data_pagamento(), "%Y-%m-%d"
                    ).date()
                    if (
                        periodo["inicio"] <= data_pag <= periodo["fim"]
                        and p.get_status().lower() == "confirmado"
                    ):
                        faturamento += Decimal(str(p.get_valor_total()))
                except:
                    continue

            total_hospedes = len(hospedes_unicos)
            ticket_medio = (
                faturamento / total_hospedes if total_hospedes > 0 else Decimal("0")
            )

            return {
                "total_reservas": len(reservas_periodo),
                "total_hospedes": total_hospedes,
                "faturamento": faturamento,
                "ticket_medio": ticket_medio,
                "reservas_por_status": dict(reservas_por_status),
                "reservas_por_tipo": dict(reservas_por_tipo),
                "reservas_detalhe": reservas_periodo,
                "consumos": consumos,
                "adicionais": adicionais,
                "periodo": periodo,
            }

        except Exception as e:
            st.error(f"Erro ao processar dados: {e}")
            return None

    @staticmethod
    def _calcular_periodo_anterior(periodo):
        """Calcula métricas do período anterior para comparação"""
        dias = (periodo["fim"] - periodo["inicio"]).days
        inicio_ant = periodo["inicio"] - timedelta(days=dias + 1)
        fim_ant = periodo["inicio"] - timedelta(days=1)

        periodo_ant = {"inicio": inicio_ant, "fim": fim_ant}
        return RelatoriosKimiUI._processar_dados_completos(periodo_ant)

    @staticmethod
    def _render_visao_geral_plus(periodo):
        """Visão geral com cards e gráficos"""
        dados = RelatoriosKimiUI._processar_dados_completos(periodo)
        if not dados:
            return

        dados_ant = RelatoriosKimiUI._calcular_periodo_anterior(periodo)

        st.subheader(
            f"{periodo['inicio'].strftime('%d/%m/%Y')} a {periodo['fim'].strftime('%d/%m/%Y')}"
        )

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            delta = RelatoriosKimiUI._calcular_variacao(
                dados["total_reservas"], dados_ant["total_reservas"] if dados_ant else 0
            )
            st.metric("Total Reservas", dados["total_reservas"], delta)

        with col2:
            delta = RelatoriosKimiUI._calcular_variacao(
                dados["total_hospedes"], dados_ant["total_hospedes"] if dados_ant else 0
            )
            st.metric("Hóspedes Únicos", dados["total_hospedes"], delta)

        with col3:
            fat_str = (
                f"R$ {float(dados['faturamento']):,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", ".")
            )
            fat_ant = float(dados_ant["faturamento"]) if dados_ant else 0
            delta = RelatoriosKimiUI._calcular_variacao(
                float(dados["faturamento"]), fat_ant
            )
            st.metric("Faturamento", fat_str, delta)

        with col4:
            tk_str = (
                f"R$ {float(dados['ticket_medio']):,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", ".")
            )
            tk_ant = float(dados_ant["ticket_medio"]) if dados_ant else 0
            delta = RelatoriosKimiUI._calcular_variacao(
                float(dados["ticket_medio"]), tk_ant
            )
            st.metric("Ticket Médio", tk_str, delta)

        st.divider()

        col_graf1, col_graf2 = st.columns(2)

        with col_graf1:
            st.markdown("#### Status das Reservas")
            if dados["reservas_por_status"]:
                df_status = pd.DataFrame(
                    [
                        {"Status": k, "Quantidade": v}
                        for k, v in dados["reservas_por_status"].items()
                    ]
                )

                if PLOTLY_AVAILABLE:
                    fig = px.pie(
                        df_status,
                        values="Quantidade",
                        names="Status",
                        hole=0.4,
                        color_discrete_sequence=px.colors.sequential.Blues,
                    )
                    fig.update_layout(height=300, showlegend=True)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.bar_chart(df_status.set_index("Status"))
            else:
                st.info("Sem dados de status")

        with col_graf2:
            st.markdown("#### Ocupação por Tipo de Quarto")
            if dados["reservas_por_tipo"]:
                df_tipo = pd.DataFrame(
                    [
                        {"Tipo": k, "Reservas": v}
                        for k, v in dados["reservas_por_tipo"].items()
                    ]
                ).sort_values("Reservas", ascending=True)

                if PLOTLY_AVAILABLE:
                    fig = px.bar(
                        df_tipo,
                        x="Reservas",
                        y="Tipo",
                        orientation="h",
                        color="Reservas",
                        color_continuous_scale="Blues",
                    )
                    fig.update_layout(height=300, showlegend=False)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.bar_chart(df_tipo.set_index("Tipo"))
            else:
                st.info("Sem dados por tipo")

    @staticmethod
    def _render_comparativo_temporal(periodo):
        """Comparativo mês a mês"""
        st.subheader("Análise Temporal")

        meses_dados = []
        hoje = periodo["fim"]

        for i in range(5, -1, -1):
            data_ref = hoje - timedelta(days=i * 30)
            mes = data_ref.month
            ano = data_ref.year

            inicio_mes = date(ano, mes, 1)
            fim_mes = date(ano, mes, calendar.monthrange(ano, mes)[1])

            dados_mes = RelatoriosKimiUI._processar_dados_completos(
                {"inicio": inicio_mes, "fim": fim_mes}
            )
            if dados_mes:
                dados_mes["label"] = f"{calendar.month_abbr[mes]}/{ano}"
                meses_dados.append(dados_mes)

        if not meses_dados:
            st.info("Dados insuficientes para análise temporal")
            return

        df_evolucao = pd.DataFrame(
            [
                {
                    "Período": d["label"],
                    "Reservas": d["total_reservas"],
                    "Faturamento": float(d["faturamento"]),
                }
                for d in meses_dados
            ]
        )

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("#### Evolução de Reservas")
            if PLOTLY_AVAILABLE:
                fig = px.line(
                    df_evolucao,
                    x="Período",
                    y="Reservas",
                    markers=True,
                    title="Reservas por Mês",
                )
                fig.update_layout(height=300)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.line_chart(df_evolucao.set_index("Período")[["Reservas"]])

        with col2:
            st.markdown("#### Evolução de Faturamento")
            if PLOTLY_AVAILABLE:
                fig = px.bar(
                    df_evolucao,
                    x="Período",
                    y="Faturamento",
                    title="Faturamento por Mês (R$)",
                    text_auto=True,
                )
                fig.update_layout(height=300)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.bar_chart(df_evolucao.set_index("Período")[["Faturamento"]])

        st.markdown("#### Tabela Comparativa (Últimos 3 meses)")
        df_comp = pd.DataFrame(
            [
                {
                    "Período": d["label"],
                    "Reservas": d["total_reservas"],
                    "Hóspedes": d["total_hospedes"],
                    "Faturamento": f"R$ {float(d['faturamento']):,.2f}".replace(
                        ",", "X"
                    )
                    .replace(".", ",")
                    .replace("X", "."),
                    "Ticket Médio": f"R$ {float(d['ticket_medio']):,.2f}".replace(
                        ",", "X"
                    )
                    .replace(".", ",")
                    .replace("X", "."),
                }
                for d in meses_dados[-3:]
            ]
        )
        st.dataframe(df_comp, hide_index=True, use_container_width=True)

    @staticmethod
    def _render_servicos_adicionais(periodo):
        """Análise de serviços com ranking"""
        st.subheader("Análise de Serviços Adicionais")

        consumos = View.consumo_listar()
        if not consumos:
            st.info("Nenhum consumo registrado")
            return

        adicionais_stats = defaultdict(
            lambda: {"qtd": 0, "total": Decimal("0"), "nome": ""}
        )
        adicionais_map = {
            a.get_id_adicional(): a for a in View.adicional_listar() or []
        }

        for c in consumos:
            try:
                data_c = datetime.strptime(c.get_data_consumo(), "%Y-%m-%d %H:%M:%S")
                if periodo["inicio"] <= data_c.date() <= periodo["fim"]:
                    id_adic = c.get_id_adicional()
                    adic = adicionais_map.get(id_adic)
                    if adic:
                        adicionais_stats[id_adic]["nome"] = adic.get_descricao()
                        adicionais_stats[id_adic]["qtd"] += c.get_quantidade()
                        valor = Decimal(str(adic.get_valor())) * Decimal(
                            c.get_quantidade()
                        )
                        adicionais_stats[id_adic]["total"] += valor
            except:
                continue

        if not adicionais_stats:
            st.info("Nenhum consumo no período selecionado")
            return

        df_adic = pd.DataFrame(
            [
                {
                    "ID": k,
                    "Adicional": v["nome"],
                    "Quantidade": v["qtd"],
                    "Total": float(v["total"]),
                }
                for k, v in adicionais_stats.items()
            ]
        ).sort_values("Total", ascending=False)

        col1, col2, col3 = st.columns(3)
        total_itens = df_adic["Quantidade"].sum()
        total_valor = df_adic["Total"].sum()

        with col1:
            st.metric("Total Itens", int(total_itens))
        with col2:
            st.metric(
                "Receita Adicionais",
                f"R$ {total_valor:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", "."),
            )
        with col3:
            ticket = total_valor / total_itens if total_itens > 0 else 0
            st.metric(
                "Ticket Médio/Item",
                f"R$ {ticket:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", "."),
            )

        st.divider()

        col_graf1, col_graf2 = st.columns(2)

        with col_graf1:
            st.markdown("#### Top por Quantidade")
            if PLOTLY_AVAILABLE:
                fig = px.bar(
                    df_adic.head(10),
                    x="Adicional",
                    y="Quantidade",
                    color="Quantidade",
                    color_continuous_scale="Reds",
                )
                fig.update_layout(height=350, xaxis_tickangle=-45)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.bar_chart(df_adic.head(10).set_index("Adicional")[["Quantidade"]])

        with col_graf2:
            st.markdown("#### Impacto Financeiro")
            if PLOTLY_AVAILABLE:
                fig = px.pie(
                    df_adic.head(8),
                    values="Total",
                    names="Adicional",
                    title="Distribuição da Receita",
                )
                fig.update_layout(height=350)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.bar_chart(df_adic.head(8).set_index("Adicional")[["Total"]])

        with st.expander("Ver Detalhamento Completo"):
            df_display = df_adic.copy()
            df_display["Total"] = df_display["Total"].apply(
                lambda x: f"R$ {x:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", ".")
            )
            st.dataframe(df_display, hide_index=True, use_container_width=True)

    @staticmethod
    def _render_exportacao_profissional(periodo):
        """Exportação multi-formato"""
        st.subheader("Exportar Relatórios")

        dados = RelatoriosKimiUI._processar_dados_completos(periodo)
        if not dados:
            st.error("Dados insuficientes para exportação")
            return

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("#### Microsoft Excel")
            if st.button("Gerar Excel", use_container_width=True, key="btn_excel"):
                try:
                    buffer = RelatoriosKimiUI._gerar_excel_profissional(dados, periodo)
                    st.success("Excel gerado!")
                    st.download_button(
                        label="Baixar .xlsx",
                        data=buffer,
                        file_name=f"inncontrol_relatorio_{periodo['inicio']}_{periodo['fim']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Erro ao gerar Excel: {e}")

        with col2:
            st.markdown("#### PDF")
            if st.button("Gerar PDF", use_container_width=True, key="btn_pdf"):
                try:
                    if REPORTLAB_AVAILABLE:
                        buffer = RelatoriosKimiUI._gerar_pdf_reportlab(dados, periodo)
                        mime = "application/pdf"
                        nome = f"inncontrol_relatorio_{periodo['inicio']}_{periodo['fim']}.pdf"
                    elif FPDF_AVAILABLE:
                        buffer = RelatoriosKimiUI._gerar_pdf_simples(dados, periodo)
                        mime = "application/pdf"
                        nome = f"inncontrol_relatorio_{periodo['inicio']}_{periodo['fim']}.pdf"
                    else:
                        st.error("Instale reportlab ou fpdf para gerar PDFs")
                        return

                    st.success("PDF gerado!")
                    st.download_button(
                        label="Baixar .pdf",
                        data=buffer,
                        file_name=nome,
                        mime=mime,
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Erro ao gerar PDF: {e}")

    @staticmethod
    def _gerar_excel_profissional(dados, periodo):
        """Excel multi-abas"""
        buffer = BytesIO()

        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl ou xlsxwriter necessário")

        if "openpyxl" in globals() or "openpyxl" in dir():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()

            ws = wb.active
            ws.title = "Resumo Executivo"

            ws["A1"] = f"Relatório InnControl - {periodo['inicio']} a {periodo['fim']}"
            ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
            ws["A1"].fill = PatternFill(
                start_color="4A90E2", end_color="4A90E2", fill_type="solid"
            )
            ws.merge_cells("A1:D1")

            metricas = [
                ("Total de Reservas", dados["total_reservas"]),
                ("Total de Hóspedes", dados["total_hospedes"]),
                ("Faturamento Total", f"R$ {float(dados['faturamento']):,.2f}"),
                ("Ticket Médio", f"R$ {float(dados['ticket_medio']):,.2f}"),
            ]

            for idx, (label, valor) in enumerate(metricas, start=3):
                ws[f"A{idx}"] = label
                ws[f"B{idx}"] = valor
                ws[f"A{idx}"].font = Font(bold=True)

            ws2 = wb.create_sheet("Status das Reservas")
            ws2.append(["Status", "Quantidade"])
            for status, qtd in dados["reservas_por_status"].items():
                ws2.append([status, qtd])

            ws3 = wb.create_sheet("Tipos de Quarto")
            ws3.append(["Tipo de Quarto", "Reservas"])
            for tipo, qtd in dados["reservas_por_tipo"].items():
                ws3.append([tipo, qtd])

            wb.save(buffer)
        else:
            import xlsxwriter

            workbook = xlsxwriter.Workbook(buffer)

            worksheet = workbook.add_worksheet("Resumo")
            worksheet.write("A1", f"Relatório InnControl")
            worksheet.write("A3", "Total Reservas")
            worksheet.write("B3", dados["total_reservas"])
            workbook.close()

        buffer.seek(0)
        return buffer

    @staticmethod
    def _gerar_pdf_reportlab(dados, periodo):
        """PDF profissional usando ReportLab"""
        buffer = BytesIO()

        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#1f4788"),
            alignment=TA_CENTER,
            spaceAfter=30,
        )

        section_style = ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#2c5aa0"),
            spaceAfter=12,
        )

        elements.append(
            Paragraph(
                f"<b>Relatório Administrativo InnControl</b><br/>{periodo['inicio']} a {periodo['fim']}",
                title_style,
            )
        )

        elements.append(Paragraph("Resumo Executivo", section_style))

        dados_tabela = [
            ["Métrica", "Valor"],
            ["Total de Reservas", str(dados["total_reservas"])],
            ["Total de Hóspedes", str(dados["total_hospedes"])],
            ["Faturamento Total", f"R$ {float(dados['faturamento']):,.2f}"],
            ["Ticket Médio", f"R$ {float(dados['ticket_medio']):,.2f}"],
        ]

        tabela = Table(dados_tabela, colWidths=[8 * cm, 6 * cm])
        tabela.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a90e2")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(tabela)
        elements.append(Spacer(1, 0.5 * cm))

        if dados["reservas_por_status"]:
            elements.append(Paragraph("Distribuição por Status", section_style))
            status_data = [["Status", "Quantidade"]]
            for status, qtd in dados["reservas_por_status"].items():
                status_data.append([status, str(qtd)])

            tabela_status = Table(status_data, colWidths=[8 * cm, 6 * cm])
            tabela_status.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a90e2")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            elements.append(tabela_status)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _gerar_pdf_simples(dados, periodo):
        """Fallback PDF usando FPDF"""
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(
            0,
            10,
            f"Relatorio InnControl - {periodo['inicio']} a {periodo['fim']}",
            ln=True,
            align="C",
        )
        pdf.ln(10)

        pdf.set_font("Arial", "", 12)
        pdf.cell(0, 8, f"Total Reservas: {dados['total_reservas']}", ln=True)
        pdf.cell(0, 8, f"Hospedes: {dados['total_hospedes']}", ln=True)
        pdf.cell(0, 8, f"Faturamento: R$ {float(dados['faturamento']):,.2f}", ln=True)

        pdf_content = pdf.output(dest="S")
        if isinstance(pdf_content, str):
            pdf_content = pdf_content.encode("latin-1", "replace")

        buffer = BytesIO(pdf_content)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _calcular_variacao(atual, anterior):
        """Calcula string de variação percentual"""
        if anterior == 0:
            return None
        var = ((atual - anterior) / anterior) * 100
        return f"{var:+.1f}%"
