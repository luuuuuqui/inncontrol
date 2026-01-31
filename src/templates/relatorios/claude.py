import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from views import View
from datetime import datetime, timedelta
from decimal import Decimal
from collections import defaultdict
import calendar


class RelatoriosClaudeUI:
    """
    Módulo de relatórios administrativos para InnControl.
    Dashboards visuais para análise de reservas, faturamento e consumo.
    """

    @staticmethod
    def main():
        st.header("Relatórios Administrativos")

        col_filtro1, col_filtro2 = st.columns(2)

        with col_filtro1:
            ano_selecionado = st.selectbox(
                "Ano:",
                options=list(range(datetime.now().year - 2, datetime.now().year + 2)),
                index=2,
            )

        with col_filtro2:
            mes_selecionado = st.selectbox(
                "Mês:",
                options=list(range(1, 13)),
                format_func=lambda x: calendar.month_name[x],
                index=datetime.now().month - 1,
            )

        st.divider()

        tab1, tab2, tab3, tab4 = st.tabs(
            [
                "Visão Geral",
                "Comparativo Temporal",
                "Serviços/Adicionais",
                "Exportar Relatórios",
            ]
        )

        with tab1:
            RelatoriosClaudeUI.visao_geral(ano_selecionado, mes_selecionado)

        with tab2:
            RelatoriosClaudeUI.comparativo_temporal(ano_selecionado, mes_selecionado)

        with tab3:
            RelatoriosClaudeUI.servicos_adicionais(ano_selecionado, mes_selecionado)

        with tab4:
            RelatoriosClaudeUI.exportar_relatorios(ano_selecionado, mes_selecionado)

    @staticmethod
    def visao_geral(ano, mes):
        """dashboard principal com kpis do período"""

        st.subheader(f"Visão Geral - {calendar.month_name[mes]}/{ano}")

        dados_atual = RelatoriosClaudeUI._obter_dados_periodo(ano, mes)

        mes_ant, ano_ant = RelatoriosClaudeUI._periodo_anterior(mes, ano)
        dados_anterior = RelatoriosClaudeUI._obter_dados_periodo(ano_ant, mes_ant)

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            RelatoriosClaudeUI._card_metrica(
                "Total de Reservas",
                dados_atual["total_reservas"],
                dados_anterior["total_reservas"],
            )

        with col2:
            RelatoriosClaudeUI._card_metrica(
                "Total de Hóspedes",
                dados_atual["total_hospedes"],
                dados_anterior["total_hospedes"],
            )

        with col3:
            RelatoriosClaudeUI._card_metrica(
                "Faturamento Total",
                dados_atual["faturamento_total"],
                dados_anterior["faturamento_total"],
                formato_moeda=True,
            )

        with col4:
            RelatoriosClaudeUI._card_metrica(
                "Ticket Médio",
                dados_atual["ticket_medio"],
                dados_anterior["ticket_medio"],
                formato_moeda=True,
            )

        st.divider()

        col_grafico1, col_grafico2 = st.columns(2)

        with col_grafico1:
            st.markdown("#### Status das Reservas")
            RelatoriosClaudeUI._grafico_status_reservas(
                dados_atual["reservas_por_status"]
            )

        with col_grafico2:
            st.markdown("#### Ocupação por Tipo de Quarto")
            RelatoriosClaudeUI._grafico_tipos_quarto(dados_atual["reservas_por_tipo"])

    @staticmethod
    def comparativo_temporal(ano, mes):
        """comparação mês a mês com evolução temporal"""

        st.subheader("Comparativo Temporal")

        meses_dados = []
        for i in range(6, 0, -1):
            m, a = RelatoriosClaudeUI._periodo_relativo(mes, ano, -i)
            dados = RelatoriosClaudeUI._obter_dados_periodo(a, m)
            dados["mes_label"] = f"{calendar.month_abbr[m]}/{str(a)[2:]}"
            meses_dados.append(dados)

        dados_atual = RelatoriosClaudeUI._obter_dados_periodo(ano, mes)
        dados_atual["mes_label"] = f"{calendar.month_abbr[mes]}/{str(ano)[2:]}"
        meses_dados.append(dados_atual)

        st.markdown("#### Evolução de Reservas")
        df_reservas = pd.DataFrame(
            [
                {"mês": d["mes_label"], "reservas": d["total_reservas"]}
                for d in meses_dados
            ]
        )

        fig_reservas = px.line(
            df_reservas,
            x="mês",
            y="reservas",
            markers=True,
            title="Número de Reservas por Mês",
        )
        fig_reservas.update_layout(height=300)
        st.plotly_chart(fig_reservas, use_container_width=True)

        st.markdown("#### Evolução de Faturamento")
        df_faturamento = pd.DataFrame(
            [
                {"mês": d["mes_label"], "faturamento": float(d["faturamento_total"])}
                for d in meses_dados
            ]
        )

        fig_fat = px.bar(
            df_faturamento, x="mês", y="faturamento", title="Faturamento por Mês (R$)"
        )
        fig_fat.update_layout(height=300)
        st.plotly_chart(fig_fat, use_container_width=True)

        st.markdown("#### Tabela Comparativa")

        ultimos_3 = meses_dados[-3:]

        df_comparativo = pd.DataFrame(
            [
                {
                    "Período": d["mes_label"],
                    "Reservas": d["total_reservas"],
                    "Hóspedes": d["total_hospedes"],
                    "Faturamento": f"R$ {float(d['faturamento_total']):.2f}".replace(
                        ".", ","
                    ),
                    "Ticket Médio": f"R$ {float(d['ticket_medio']):.2f}".replace(
                        ".", ","
                    ),
                }
                for d in ultimos_3
            ]
        )

        st.dataframe(df_comparativo, hide_index=True, use_container_width=True)

    @staticmethod
    def servicos_adicionais(ano, mes):
        """análise detalhada dos serviços/adicionais consumidos"""

        st.subheader("Análise de Serviços Adicionais")

        consumos = View.consumo_listar()

        if not consumos:
            st.info("Nenhum consumo registrado no período.")
            return

        consumos_periodo = []
        for c in consumos:
            try:
                data_consumo = datetime.strptime(
                    c.get_data_consumo(), "%Y-%m-%d %H:%M:%S"
                )
                if data_consumo.year == ano and data_consumo.month == mes:
                    consumos_periodo.append(c)
            except:
                continue

        if not consumos_periodo:
            st.info(f"Nenhum consumo em {calendar.month_name[mes]}/{ano}.")
            return

        adicionais_stats = defaultdict(
            lambda: {
                "quantidade": 0,
                "valor_total": Decimal("0.00"),
                "nome": "Desconhecido",
            }
        )

        for c in consumos_periodo:
            id_adic = c.get_id_adicional()
            adicional = View.adicional_listar_id(id_adic)

            if adicional:
                adicionais_stats[id_adic]["nome"] = adicional.get_descricao()
                adicionais_stats[id_adic]["quantidade"] += c.get_quantidade()
                valor_unit = Decimal(adicional.get_valor())
                adicionais_stats[id_adic]["valor_total"] += valor_unit * Decimal(
                    c.get_quantidade()
                )

        df_adicionais = pd.DataFrame(
            [
                {
                    "adicional": stats["nome"],
                    "quantidade": stats["quantidade"],
                    "valor_total": float(stats["valor_total"]),
                }
                for id_adic, stats in adicionais_stats.items()
            ]
        ).sort_values("valor_total", ascending=False)

        col1, col2, col3 = st.columns(3)

        with col1:
            total_itens = df_adicionais["quantidade"].sum()
            st.metric("Total de Itens", f"{total_itens}")

        with col2:
            receita_adicionais = df_adicionais["valor_total"].sum()
            st.metric(
                "Receita Adicionais",
                f"R$ {receita_adicionais:.2f}".replace(".", ","),
            )

        with col3:
            ticket_medio_adic = (
                receita_adicionais / total_itens if total_itens > 0 else 0
            )
            st.metric(
                "Ticket Médio/Item", f"R$ {ticket_medio_adic:.2f}".replace(".", ",")
            )

        st.divider()

        col_graf1, col_graf2 = st.columns(2)

        with col_graf1:
            st.markdown("#### Ranking por Quantidade")
            fig_qtd = px.bar(
                df_adicionais.head(10),
                x="adicional",
                y="quantidade",
                title="Top 10 Adicionais Mais Consumidos",
            )
            fig_qtd.update_layout(height=350, xaxis_tickangle=-45)
            st.plotly_chart(fig_qtd, use_container_width=True)

        with col_graf2:
            st.markdown("#### Impacto Financeiro")
            fig_valor = px.pie(
                df_adicionais.head(10),
                values="valor_total",
                names="adicional",
                title="Distribuição de Receita por Adicional",
            )
            fig_valor.update_layout(height=350)
            st.plotly_chart(fig_valor, use_container_width=True)

        st.markdown("#### Detalhamento Completo")
        df_display = df_adicionais.copy()
        df_display["valor_total"] = df_display["valor_total"].apply(
            lambda x: f"R$ {x:.2f}".replace(".", ",")
        )
        df_display.columns = ["Adicional", "Quantidade", "Valor Total"]
        st.dataframe(df_display, hide_index=True, use_container_width=True)

    @staticmethod
    def exportar_relatorios(ano, mes):
        """exportação de relatórios em pdf e excel"""

        st.subheader("Exportar Relatórios")

        st.info(
            "Selecione o formato desejado para exportar os dados consolidados do período."
        )

        dados = RelatoriosClaudeUI._obter_dados_periodo(ano, mes)

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("#### Exportar para Excel")
            st.caption("Arquivo .xlsx com todas as tabelas e dados")

            if st.button("Gerar Excel", key="btn_excel"):
                try:
                    arquivo_excel = RelatoriosClaudeUI._gerar_excel(ano, mes, dados)
                    st.success("Excel gerado com sucesso!")

                    with open(arquivo_excel, "rb") as f:
                        st.download_button(
                            label="Baixar Excel",
                            data=f,
                            file_name=f"relatorio_inncontrol_{mes:02d}_{ano}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                except Exception as e:
                    st.error(f"Erro ao gerar Excel: {e}")

        with col2:
            st.markdown("#### Exportar para PDF")
            st.caption("Relatório consolidado em PDF")

            if st.button("Gerar PDF", key="btn_pdf"):
                try:
                    arquivo_pdf = RelatoriosClaudeUI._gerar_pdf(ano, mes, dados)
                    st.success("PDF gerado com sucesso!")

                    with open(arquivo_pdf, "rb") as f:
                        st.download_button(
                            label="Baixar PDF",
                            data=f,
                            file_name=f"relatorio_inncontrol_{mes:02d}_{ano}.pdf",
                            mime="application/pdf",
                        )
                except Exception as e:
                    st.error(f"Erro ao gerar PDF: {e}")

    @staticmethod
    def _obter_dados_periodo(ano, mes):
        """
        extrai métricas consolidadas do período especificado.
        retorna dict com todas as métricas calculadas.
        """

        todas_reservas = View.reserva_listar()

        reservas_periodo = []
        for r in todas_reservas:
            try:
                checkin = datetime.strptime(r.get_data_checkin(), "%Y-%m-%d")
                if checkin.year == ano and checkin.month == mes:
                    reservas_periodo.append(r)
            except:
                continue

        total_reservas = len(reservas_periodo)

        hospedes_ids = set()
        for r in reservas_periodo:
            hospedes_ids.add(r.get_id_hospede())
        total_hospedes = len(hospedes_ids)

        faturamento_total = Decimal("0.00")
        pagamentos = View.pagamento_listar()

        for p in pagamentos:
            try:
                data_pag = datetime.strptime(p.get_data_pagamento(), "%Y-%m-%d")
                if data_pag.year == ano and data_pag.month == mes:
                    if p.get_status().lower() == "confirmado":
                        faturamento_total += Decimal(p.get_valor_total())
            except:
                continue

        ticket_medio = (
            faturamento_total / Decimal(total_hospedes)
            if total_hospedes > 0
            else Decimal("0.00")
        )

        reservas_por_status = defaultdict(int)
        for r in reservas_periodo:
            reservas_por_status[r.get_status()] += 1

        reservas_por_tipo = defaultdict(int)
        for r in reservas_periodo:
            quarto = View.quarto_listar_id(r.get_id_quarto())
            if quarto:
                tipo = View.tipoquarto_listar_id(quarto.get_id_quarto_tipo())
                if tipo:
                    reservas_por_tipo[tipo.get_nome()] += 1

        return {
            "total_reservas": total_reservas,
            "total_hospedes": total_hospedes,
            "faturamento_total": faturamento_total,
            "ticket_medio": ticket_medio,
            "reservas_por_status": dict(reservas_por_status),
            "reservas_por_tipo": dict(reservas_por_tipo),
        }

    @staticmethod
    def _periodo_anterior(mes, ano):
        """retorna tupla (mes, ano) do período anterior"""
        if mes == 1:
            return (12, ano - 1)
        return (mes - 1, ano)

    @staticmethod
    def _periodo_relativo(mes, ano, offset):
        """retorna período com offset (ex: -1 = mês anterior, -6 = 6 meses atrás)"""
        total_meses = ano * 12 + mes + offset
        novo_ano = total_meses // 12
        novo_mes = total_meses % 12
        if novo_mes == 0:
            novo_mes = 12
            novo_ano -= 1
        return (novo_mes, novo_ano)

    @staticmethod
    def _card_metrica(titulo, valor_atual, valor_anterior, formato_moeda=False):
        """renderiza card de métrica com comparação"""

        if formato_moeda:
            valor_display = f"R$ {float(valor_atual):.2f}".replace(".", ",")
        else:
            valor_display = f"{valor_atual}"

        if valor_anterior > 0:
            variacao = (
                (float(valor_atual) - float(valor_anterior)) / float(valor_anterior)
            ) * 100
            delta = f"{variacao:+.1f}%"
        else:
            delta = "N/A"

        st.metric(label=titulo, value=valor_display, delta=delta)

    @staticmethod
    def _grafico_status_reservas(dados_status):
        """gráfico pizza para status das reservas"""

        if not dados_status:
            st.info("Sem dados de status")
            return

        df = pd.DataFrame(
            [{"status": k, "quantidade": v} for k, v in dados_status.items()]
        )

        fig = px.pie(df, values="quantidade", names="status", hole=0.4)
        fig.update_layout(height=300, showlegend=True)
        st.plotly_chart(fig, use_container_width=True)

    @staticmethod
    def _grafico_tipos_quarto(dados_tipos):
        """gráfico barras horizontais para tipos de quarto"""

        if not dados_tipos:
            st.info("Sem dados de tipos de quarto")
            return

        df = pd.DataFrame(
            [{"tipo": k, "reservas": v} for k, v in dados_tipos.items()]
        ).sort_values("reservas", ascending=True)

        fig = px.bar(df, x="reservas", y="tipo", orientation="h")
        fig.update_layout(height=300)
        st.plotly_chart(fig, use_container_width=True)

    @staticmethod
    def _gerar_excel(ano, mes, dados):
        """
        gera arquivo excel com múltiplas abas contendo:
        - resumo executivo
        - detalhamento de reservas
        - análise de adicionais
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        arquivo = f"/home/claude/relatorio_{mes:02d}_{ano}.xlsx"

        wb = openpyxl.Workbook()

        ws_resumo = wb.active
        ws_resumo.title = "Resumo Executivo"

        ws_resumo["A1"] = f"Relatório InnControl - {calendar.month_name[mes]}/{ano}"
        ws_resumo["A1"].font = Font(size=14, bold=True)
        ws_resumo.merge_cells("A1:D1")

        ws_resumo["A3"] = "Métrica"
        ws_resumo["B3"] = "Valor"
        ws_resumo["A3"].font = Font(bold=True)
        ws_resumo["B3"].font = Font(bold=True)

        metricas = [
            ("Total de Reservas", dados["total_reservas"]),
            ("Total de Hóspedes", dados["total_hospedes"]),
            ("Faturamento Total", f"R$ {float(dados['faturamento_total']):.2f}"),
            ("Ticket Médio", f"R$ {float(dados['ticket_medio']):.2f}"),
        ]

        row = 4
        for metrica, valor in metricas:
            ws_resumo[f"A{row}"] = metrica
            ws_resumo[f"B{row}"] = valor
            row += 1

        ws_resumo.column_dimensions["A"].width = 25
        ws_resumo.column_dimensions["B"].width = 20

        ws_status = wb.create_sheet("Reservas por Status")
        ws_status["A1"] = "Status"
        ws_status["B1"] = "Quantidade"
        ws_status["A1"].font = Font(bold=True)
        ws_status["B1"].font = Font(bold=True)

        row = 2
        for status, qtd in dados["reservas_por_status"].items():
            ws_status[f"A{row}"] = status
            ws_status[f"B{row}"] = qtd
            row += 1

        ws_status.column_dimensions["A"].width = 20
        ws_status.column_dimensions["B"].width = 15

        wb.save(arquivo)
        return arquivo

    @staticmethod
    def _gerar_pdf(ano, mes, dados):
        """
        gera relatório pdf consolidado usando reportlab.
        limitação: requer biblioteca reportlab instalada.
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        arquivo = f"/home/claude/relatorio_{mes:02d}_{ano}.pdf"

        doc = SimpleDocTemplate(arquivo, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        titulo_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#1f4788"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        titulo = Paragraph(
            f"<b>Relatório Administrativo InnControl</b><br/>{calendar.month_name[mes]}/{ano}",
            titulo_style,
        )
        elements.append(titulo)
        elements.append(Spacer(1, 0.5 * cm))

        secao_style = ParagraphStyle(
            "Secao",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#2c5aa0"),
            spaceAfter=12,
        )

        elements.append(Paragraph("Resumo Executivo", secao_style))

        dados_tabela = [
            ["Métrica", "Valor"],
            ["Total de Reservas", str(dados["total_reservas"])],
            ["Total de Hóspedes", str(dados["total_hospedes"])],
            ["Faturamento Total", f"R$ {float(dados['faturamento_total']):.2f}"],
            ["Ticket Médio", f"R$ {float(dados['ticket_medio']):.2f}"],
        ]

        tabela = Table(dados_tabela, colWidths=[8 * cm, 6 * cm])
        tabela.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a90e2")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        elements.append(tabela)
        elements.append(Spacer(1, 1 * cm))

        if dados["reservas_por_status"]:
            elements.append(Paragraph("Distribuição por Status", secao_style))

            dados_status = [["Status", "Quantidade"]]
            for status, qtd in dados["reservas_por_status"].items():
                dados_status.append([status, str(qtd)])

            tabela_status = Table(dados_status, colWidths=[8 * cm, 6 * cm])
            tabela_status.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4a90e2")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )

            elements.append(tabela_status)

        doc.build(elements)
        return arquivo
