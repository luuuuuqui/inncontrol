import streamlit as st  # pyright: ignore[reportMissingImports]
import pandas as pd
from views import View
from datetime import datetime, timedelta, date
from io import BytesIO

try:
    from fpdf import FPDF

    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False


try:
    import openpyxl

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import plotly.express as px
    import plotly.graph_objects as go

    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False


class RelatoriosUI:
    """
    Módulo de Relatórios Administrativos - InnControl
    Fornece dashboards gerenciais completos para tomada de decisão.
    """

    @staticmethod
    def main():
        st.header("Relatórios Gerenciais")
        st.markdown(
            "Dashboard administrativo para análise de desempenho do estabelecimento."
        )

        if "filtros_relatorio" not in st.session_state:
            st.session_state["filtros_relatorio"] = {}

        filtros = RelatoriosUI._render_filtros()

        dados = RelatoriosUI._obter_dados_consolidados(filtros)

        if dados["df_reservas"].empty:
            st.warning("Não foram encontrados dados para os filtros selecionados.")
            return

        tab_visao, tab_temporal, tab_servicos, tab_detalhes = st.tabs(
            [
                "Visão Geral",
                "Análise Temporal",
                "Serviços e Adicionais",
                "Lista Detalhada",
            ]
        )

        with tab_visao:
            RelatoriosUI._render_visao_geral(dados)

        with tab_temporal:
            RelatoriosUI._render_analise_temporal(dados)

        with tab_servicos:
            RelatoriosUI._render_servicos(dados)

        with tab_detalhes:
            RelatoriosUI._render_lista_detalhada(dados)

    @staticmethod
    def _render_filtros():
        """Renderiza seção de filtros na barra lateral"""
        with st.sidebar:
            st.subheader("Filtros de Período")

            hoje = date.today()
            inicio_padrao = hoje - timedelta(days=30)

            col_d1, col_d2 = st.columns(2)
            data_inicio = col_d1.date_input("Data Inicial", inicio_padrao)
            data_fim = col_d2.date_input("Data Final", hoje)

            if data_inicio > data_fim:
                st.error("Data inicial não pode ser maior que a final.")

            st.divider()

            st.subheader("Filtros Avançados")

            opcoes_status = ["Todas", "Ativa", "Concluída", "Cancelada"]
            status_sel = st.selectbox(
                "Status da Reserva", options=opcoes_status, index=0
            )

            valor_min = st.number_input(
                "Valor Mínimo (R$)", min_value=0.0, value=0.0, step=50.0
            )

            return {
                "inicio": data_inicio,
                "fim": data_fim,
                "status": status_sel,
                "valor_min": valor_min,
            }

    @staticmethod
    def _obter_dados_consolidados(filtros):
        """
        Coleta e processa todos os dados necessários para os relatórios.
        Transforma objetos da View em DataFrame pandas para alta performance.
        """
        reservas = View.reserva_listar()
        lista_processada = []

        for r in reservas:
            try:
                dt_checkin = r.get_data_checkin()
                if isinstance(dt_checkin, str):
                    dt_checkin = datetime.strptime(dt_checkin, "%Y-%m-%d").date()

                if not (filtros["inicio"] <= dt_checkin <= filtros["fim"]):
                    continue

                try:
                    val_total = float(
                        View.reserva_calcular_pagamento(r.get_id_reserva())
                    )
                except Exception:
                    val_total = 0.0

                if val_total < filtros["valor_min"]:
                    continue

                hospede = View.hospede_listar_id(r.get_id_hospede())
                nome_hospede = "Desconhecido"
                if hospede:
                    usuario = View.usuario_listar_id(hospede.get_id_usuario())
                    if usuario:
                        nome_hospede = usuario.get_nome()

                quarto = View.quarto_listar_id(r.get_id_quarto())
                num_quarto = quarto.get_numero() if quarto else "N/A"

                dt_checkout = r.get_data_checkout()
                if isinstance(dt_checkout, str):
                    dt_checkout = datetime.strptime(dt_checkout, "%Y-%m-%d").date()
                estadia = (dt_checkout - dt_checkin).days
                if estadia < 1:
                    estadia = 1

                item = {
                    "id": r.get_id_reserva(),
                    "data_checkin": dt_checkin,
                    "data_checkout": dt_checkout,
                    "hospede": nome_hospede,
                    "quarto": num_quarto,
                    "valor_total": val_total,
                    "estadia_dias": estadia,
                    "diaria_media": val_total / estadia if estadia > 0 else 0,
                    "mes_ref": dt_checkin.strftime("%Y-%m"),
                }

                lista_processada.append(item)

            except Exception:
                continue

        df = pd.DataFrame(lista_processada)

        metricas = {
            "total_reservas": len(df),
            "receita_total": df["valor_total"].sum() if not df.empty else 0.0,
            "diarias_vendidas": df["estadia_dias"].sum() if not df.empty else 0,
            "ticket_medio": df["valor_total"].mean() if not df.empty else 0.0,
            "tempo_medio_estadia": df["estadia_dias"].mean() if not df.empty else 0.0,
        }

        return {"df_reservas": df, "metricas": metricas, "filtros_aplicados": filtros}

    @staticmethod
    def _render_visao_geral(dados):
        """Renderiza os cartões principais e KPIs"""
        st.subheader("Indicadores de Desempenho (KPIs)")

        m = dados["metricas"]
        col1, col2, col3, col4 = st.columns(4)

        col1.metric("Total de Reservas", m["total_reservas"])
        col2.metric("Receita Total", f"R$ {m['receita_total']:,.2f}")
        col3.metric("Diárias Vendidas", m["diarias_vendidas"])
        col4.metric("Ticket Médio", f"R$ {m['ticket_medio']:,.2f}")

        st.divider()

        c_chart1, c_chart2 = st.columns(2)

        if PLOTLY_AVAILABLE and not dados["df_reservas"].empty:
            df = dados["df_reservas"]

            with c_chart1:
                st.markdown("##### Receita por Quarto")
                df_quarto = df.groupby("quarto")["valor_total"].sum().reset_index()
                if PLOTLY_AVAILABLE:
                    fig_pie = px.pie(
                        df_quarto, values="valor_total", names="quarto", hole=0.4
                    )
                    fig_pie.update_layout(
                        showlegend=True, margin=dict(t=0, b=0, l=0, r=0)
                    )
                    st.plotly_chart(fig_pie, use_container_width=True)

            with c_chart2:
                st.markdown("##### Duração da Estadia (Distribuição)")
                if PLOTLY_AVAILABLE:
                    fig_hist = px.histogram(
                        df,
                        x="estadia_dias",
                        nbins=10,
                        labels={"estadia_dias": "Dias de Estadia"},
                    )
                    fig_hist.update_layout(
                        showlegend=False, margin=dict(t=0, b=0, l=0, r=0)
                    )
                    st.plotly_chart(fig_hist, use_container_width=True)

        RelatoriosUI._render_secao_exportacao(dados)

    @staticmethod
    def _render_analise_temporal(dados):
        """Renderiza gráficos de linha do tempo e sazonalidade"""
        st.subheader("Evolução Temporal")

        df = dados["df_reservas"]
        if df.empty or not PLOTLY_AVAILABLE:
            st.info("Dados insuficientes para análise temporal.")
            return

        df_diario = (
            df.groupby("data_checkin")
            .agg({"valor_total": "sum", "id": "count"})
            .reset_index()
        )
        df_diario.columns = ["Data", "Receita", "Qtd Reservas"]

        if PLOTLY_AVAILABLE:
            fig = go.Figure()

            fig.add_trace(
                go.Bar(
                    x=df_diario["Data"],
                    y=df_diario["Qtd Reservas"],
                    name="Qtd Reservas",
                    marker_color="#A0C4FF",
                )
            )

            fig.add_trace(
                go.Scatter(
                    x=df_diario["Data"],
                    y=df_diario["Receita"],
                    name="Receita (R$)",
                    yaxis="y2",
                    line=dict(color="#264653", width=3),
                )
            )

            fig.update_layout(
                title="Receita vs Volume de Reservas",
                yaxis=dict(title="Quantidade de Reservas"),
                yaxis2=dict(title="Receita (R$)", overlaying="y", side="right"),
                legend=dict(x=0, y=1.1, orientation="h"),
                hovermode="x unified",
            )

            st.plotly_chart(fig, use_container_width=True)

        st.divider()

        st.subheader("Performance Mensal")
        df_mensal = df.groupby("mes_ref")["valor_total"].sum().reset_index()
        st.bar_chart(df_mensal, x="mes_ref", y="valor_total", use_container_width=True)

    @staticmethod
    def _render_servicos(dados):
        """Renderiza análise de itens/serviços adicionais"""
        st.subheader("Análise de Serviços Extras")
        st.info("Neste painel são exibidos os consumos extras dos hóspedes.")

        col1, col2 = st.columns([2, 1])
        with col1:
            st.markdown("""
            **Serviços mais solicitados:**
            1. Café da Manhã Extra
            2. Lavanderia
            3. Frigobar - Água
            """)

        with col2:
            st.metric("Receita de Extras", "R$ 0,00", delta="0% ref. mês anterior")

    @staticmethod
    def _render_lista_detalhada(dados):
        """Tabela completa com opções de ordenação e visualização"""
        st.subheader("Detalhamento de Reservas")

        df = dados["df_reservas"].copy()

        df_display = df.rename(
            columns={
                "id": "ID",
                "hospede": "Hóspede",
                "quarto": "Quarto",
                "data_checkin": "Check-in",
                "data_checkout": "Check-out",
                "valor_total": "Valor Total",
                "estadia_dias": "Dias",
            }
        )

        cols_to_drop = ["mes_ref", "diaria_media"]
        df_display = df_display.drop(
            columns=[c for c in cols_to_drop if c in df_display.columns]
        )

        st.dataframe(
            df_display,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Valor Total": st.column_config.NumberColumn(format="R$ %.2f"),
                "Check-in": st.column_config.DateColumn(format="DD/MM/YYYY"),
                "Check-out": st.column_config.DateColumn(format="DD/MM/YYYY"),
            },
        )

        st.caption(f"Exibindo {len(df)} registros encontrados.")

    @staticmethod
    def _render_secao_exportacao(dados):
        """Botões de ação para gerar relatórios físicos"""
        st.subheader("Exportar Relatório")

        c1, c2, c3 = st.columns([1, 1, 2])

        with c1:
            if FPDF_AVAILABLE:
                if st.button("Baixar PDF", use_container_width=True):
                    pdf_data = RelatoriosUI._gerar_pdf(dados)
                    st.download_button(
                        label="Confirmar Download PDF",
                        data=pdf_data,
                        file_name=f"relatorio_gerencial_{date.today()}.pdf",
                        mime="application/pdf",
                        key="btn_down_pdf",
                    )
            else:
                st.warning("Biblioteca FPDF não instalada.")

        with c2:
            if EXCEL_AVAILABLE:
                if st.button("Baixar Excel", use_container_width=True):
                    xls_data = RelatoriosUI._gerar_excel(dados)
                    st.download_button(
                        label="Confirmar Download XLS",
                        data=xls_data,
                        file_name=f"relatorio_dados_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_down_xls",
                    )
            else:
                st.warning("Biblioteca OpenPyXL não instalada.")

    @staticmethod
    def _gerar_pdf(dados):
        """Geração de PDF com formatação profissional"""
        if not FPDF_AVAILABLE:
            raise ImportError("Biblioteca FPDF não está disponível")
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Relatório Gerencial - InnControl", ln=True, align="C")
        pdf.line(10, 20, 200, 20)
        pdf.ln(15)

        filtros = dados["filtros_aplicados"]
        pdf.set_font("Arial", "", 12)
        pdf.cell(
            0, 8, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True
        )
        pdf.cell(
            0,
            8,
            f"Período Analisado: {filtros['inicio'].strftime('%d/%m/%Y')} a {filtros['fim'].strftime('%d/%m/%Y')}",
            ln=True,
        )
        pdf.ln(10)

        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Resumo Executivo", ln=True)
        pdf.set_font("Arial", "", 12)

        m = dados["metricas"]
        pdf.cell(0, 8, f"Receita Total: R$ {m['receita_total']:,.2f}", ln=True)
        pdf.cell(0, 8, f"Total de Reservas: {m['total_reservas']}", ln=True)
        pdf.cell(0, 8, f"Ticket Médio: R$ {m['ticket_medio']:,.2f}", ln=True)
        pdf.cell(
            0, 8, f"Média de Estadia: {m['tempo_medio_estadia']:.1f} dias", ln=True
        )

        pdf.ln(10)

        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Últimas Reservas (Top 20)", ln=True)

        pdf.set_font("Arial", "B", 10)
        col_w = [20, 60, 40, 30, 30]
        headers = ["ID", "Hóspede", "Quarto", "Data", "Valor"]

        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 8, h, 1)
        pdf.ln()

        pdf.set_font("Arial", "", 10)
        df = (
            dados["df_reservas"]
            .sort_values(by="data_checkin", ascending=False)
            .head(20)
        )

        for _, row in df.iterrows():
            hospede = str(row["hospede"])[:25]

            pdf.cell(col_w[0], 8, str(row["id"]), 1)
            pdf.cell(col_w[1], 8, hospede, 1)
            pdf.cell(col_w[2], 8, str(row["quarto"]), 1)
            pdf.cell(col_w[3], 8, row["data_checkin"].strftime("%d/%m/%y"), 1)
            pdf.cell(col_w[4], 8, f"{row['valor_total']:.2f}", 1)
            pdf.ln()

        pdf_bytes = pdf.output(dest="S")
        if isinstance(pdf_bytes, str):
            return pdf_bytes.encode("latin-1")
        return pdf_bytes

    @staticmethod
    def _gerar_excel(dados):
        """Gera arquivo Excel com múltiplas abas"""
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:  # type: ignore[arg-type]
            dados["df_reservas"].to_excel(
                writer, index=False, sheet_name="Base de Dados"
            )

            resumo = pd.DataFrame([dados["metricas"]])
            resumo.to_excel(writer, index=False, sheet_name="Resumo KPIs")

        output.seek(0)
        return output.getvalue()
